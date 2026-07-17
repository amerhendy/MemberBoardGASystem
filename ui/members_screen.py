# -*- coding: utf-8 -*-
"""
members_screen.py
شاشة إدارة الأعضاء المطورة - Premium Members Management Screen
تم التخلص من التقسيم الثنائي واستخدام نافذة منبثقة ذكية (Dialog) لإضافة وتعديل بيانات الأعضاء
مع دعم كامل للحقول الديناميكية والمستندات المرفقة بتصميم فخم ومريح للعين.
"""
from PyQt6.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QFormLayout, QTableWidget, QTableWidgetItem,
    QLineEdit, QComboBox, QDateEdit, QTextEdit, QPushButton, QLabel, QFileDialog,
    QMessageBox, QGroupBox, QScrollArea, QDialog, QHeaderView,QGridLayout
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor

from config import MEMBER_TYPES, MEMBER_TYPE_LABELS_AR, MEMBER_STATUSES, MEMBER_STATUS_LABELS_AR, DOC_TYPES


# ==============================================================================
# 1. نافذة منبثقة ذكية وموسعة لإضافة أو تعديل بيانات عضو
# ==============================================================================
class MemberFormDialog(QDialog):
    def __init__(self, parent, member_id=None):
        super().__init__(parent)
        self.screen = parent
        self.member_id = member_id  # لو كان ممرراً، يعني أننا في وضع التعديل
        self.dynamic_field_widgets = {}

        if self.member_id:
            self.setWindowTitle("تعديل بيانات العضو الكريم")
        else:
            self.setWindowTitle("إضافة عضو جديد للنظام")

        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.resize(550, 650)
        self._build_ui()
        self._load_data_if_edit()

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # استخدام منطقة تمرير داخل النافذة لضمان ظهور كافة الحقول والمرفقات بسلاسة
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_widget = QWidget()
        form_layout = QVBoxLayout(scroll_widget)
        form_layout.setSpacing(15)

        # ---------- 1. بيانات العضو الأساسية ----------
        base_box = QGroupBox("البيانات الأساسية")
        base_form = QFormLayout()
        base_form.setSpacing(10)

        self.full_name_input = QLineEdit()
        self.full_name_input.setMinimumHeight(35)
        self.full_name_input.setPlaceholderText("أدخل الاسم الكامل للعضو...")

        self.member_type_input = QComboBox()
        self.member_type_input.setMinimumHeight(35)
        for t in MEMBER_TYPES:
            self.member_type_input.addItem(MEMBER_TYPE_LABELS_AR[t], t)

        self.organization_input = QLineEdit()
        self.organization_input.setMinimumHeight(35)

        self.start_date_input = QDateEdit(calendarPopup=True)
        self.start_date_input.setMinimumHeight(35)
        self.start_date_input.setDate(QDate.currentDate())

        self.end_date_input = QDateEdit(calendarPopup=True)
        self.end_date_input.setMinimumHeight(35)
        self.end_date_input.setDate(QDate.currentDate())
        self.end_date_input.setSpecialValueText(" ")

        self.status_input = QComboBox()
        self.status_input.setMinimumHeight(35)
        for s in MEMBER_STATUSES:
            self.status_input.addItem(MEMBER_STATUS_LABELS_AR[s], s)

        self.notes_input = QTextEdit()
        self.notes_input.setMaximumHeight(60)

        base_form.addRow("الاسم الكامل:", self.full_name_input)
        base_form.addRow("النوع:", self.member_type_input)
        base_form.addRow("الجهة/المؤسسة:", self.organization_input)
        base_form.addRow("تاريخ التعيين:", self.start_date_input)
        base_form.addRow("تاريخ الانتهاء (إن وجد):", self.end_date_input)
        base_form.addRow("الحالة:", self.status_input)
        base_form.addRow("ملاحظات:", self.notes_input)
        base_box.setLayout(base_form)
        form_layout.addWidget(base_box)

        # ---------- 2. الحقول الديناميكية الإضافية ----------
        self.dynamic_box = QGroupBox("حقول إضافية")
        self.dynamic_form_layout = QFormLayout()
        self.dynamic_form_layout.setSpacing(10)
        self.dynamic_box.setLayout(self.dynamic_form_layout)
        form_layout.addWidget(self.dynamic_box)

        # ربط تغير نوع العضو بإعادة بناء الحقول الديناميكية فوراً
        self.member_type_input.currentIndexChanged.connect(self._rebuild_dynamic_fields)

        # ---------- 3. المستندات المرفقة ----------
        docs_box = QGroupBox("المستندات المرفقة")
        docs_layout = QVBoxLayout()
        docs_layout.setSpacing(8)

        upload_bar = QHBoxLayout()
        self.doc_type_combo = QComboBox()
        self.doc_type_combo.setMinimumHeight(35)
        self.doc_type_combo.addItems(DOC_TYPES)
        
        upload_btn = QPushButton("رفع مستند جديد")
        upload_btn.setMinimumHeight(35)
        upload_btn.setStyleSheet("""
            QPushButton { background-color: #34495e; color: white; font-weight: bold; border-radius: 4px; padding: 0 12px; }
            QPushButton:hover { background-color: #2c3e50; }
        """)
        upload_btn.clicked.connect(self._upload_document)
        
        upload_bar.addWidget(self.doc_type_combo, 2)
        upload_bar.addWidget(upload_btn, 1)
        docs_layout.addLayout(upload_bar)

        self.documents_table = QTableWidget(0, 2)
        self.documents_table.setHorizontalHeaderLabels(["نوع المستند", "اسم الملف"])
        self.documents_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.documents_table.setMinimumHeight(120)
        self.documents_table.setStyleSheet("""
            QTableWidget { background-color: #2f3640; color: #f5f6fa; border-radius: 4px; }
            QHeaderView::section { background-color: #1e272e; color: #f5f6fa; }
        """)
        docs_layout.addWidget(self.documents_table)
        docs_box.setLayout(docs_layout)
        form_layout.addWidget(docs_box)

        scroll_area.setWidget(scroll_widget)
        main_layout.addWidget(scroll_area)

        # ---------- 4. أزرار التحكم السفلية للنافذة ----------
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)

        save_btn = QPushButton("حفظ البيانات")
        save_btn.setMinimumHeight(38)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; font-weight: bold; border: none; border-radius: 4px; }
            QPushButton:hover { background-color: #219653; }
        """)
        save_btn.clicked.connect(self._save_member)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setMinimumHeight(38)
        cancel_btn.setStyleSheet("""
            QPushButton { background-color: #7f8c8d; color: white; border: none; border-radius: 4px; }
            QPushButton:hover { background-color: #95a5a6; }
        """)
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)
        main_layout.addLayout(buttons_layout)

        # بناء حقول أولية بالاعتماد على الفئة المختارة تلقائياً
        self._rebuild_dynamic_fields()

    def _rebuild_dynamic_fields(self):
        while self.dynamic_form_layout.count():
            item = self.dynamic_form_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        self.dynamic_field_widgets = {}

        member_type = self.member_type_input.currentData()
        field_defs = self.screen.dynamic_fields_service.list_fields(applies_to=member_type)

        existing_values = {}
        if self.member_id:
            for v in self.screen.dynamic_fields_service.get_member_values(self.member_id, member_type):
                existing_values[v["field_id"]] = v["value"]

        for field in field_defs:
            widget = QLineEdit()
            widget.setMinimumHeight(33)
            field_id = field.get("id")
            value = existing_values.get(field_id, "")
            widget.setText(value or "")
            self.dynamic_form_layout.addRow(field["label_ar"] + ":", widget)
            self.dynamic_field_widgets[field_id] = widget

    def _load_data_if_edit(self):
        if not self.member_id:
            return
        member = self.screen.member_service.get_member(self.member_id)
        if not member:
            return

        self.full_name_input.setText(member["full_name"])
        idx = self.member_type_input.findData(member["member_type"])
        self.member_type_input.setCurrentIndex(max(idx, 0))
        self.organization_input.setText(member["organization"] or "")
        self.start_date_input.setDate(QDate.fromString(member["start_date"], "yyyy-MM-dd"))
        if member["end_date"]:
            self.end_date_input.setDate(QDate.fromString(member["end_date"], "yyyy-MM-dd"))
        status_idx = self.status_input.findData(member["status"])
        self.status_input.setCurrentIndex(max(status_idx, 0))
        self.notes_input.setPlainText(member["notes"] or "")

        self._rebuild_dynamic_fields()
        self._load_documents()

    def _load_documents(self):
        self.documents_table.setRowCount(0)
        if not self.member_id:
            return
        docs = self.screen.document_service.list_documents(self.member_id)
        self.documents_table.setRowCount(len(docs))
        for row, doc in enumerate(docs):
            self.documents_table.setItem(row, 0, QTableWidgetItem(doc['doc_type']))
            self.documents_table.setItem(row, 1, QTableWidgetItem(doc['original_name']))

    def _upload_document(self):
        if not self.member_id:
            QMessageBox.warning(self, "تنبيه", "الرجاء حفظ بيانات العضو أولاً قبل رفع المستندات.")
            return
        file_path, _ = QFileDialog.getOpenFileName(self, "اختر ملفًا")
        if not file_path:
            return
        member_type = self.member_type_input.currentData()
        doc_type = self.doc_type_combo.currentText()
        try:
            self.screen.document_service.add_document(self.member_id, member_type, doc_type, file_path)
            self._load_documents()
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"تعذّر رفع الملف: {e}")

    def _save_member(self):
        full_name = self.full_name_input.text().strip()
        if not full_name:
            QMessageBox.warning(self, "تنبيه", "الرجاء إدخال اسم العضو.")
            return

        member_type = self.member_type_input.currentData()
        organization = self.organization_input.text().strip()
        start_date = self.start_date_input.date().toString("yyyy-MM-dd")
        end_date = self.end_date_input.date().toString("yyyy-MM-dd")
        status = self.status_input.currentData()
        notes = self.notes_input.toPlainText().strip()

        if self.member_id:
            self.screen.member_service.update_member(
                self.member_id, full_name=full_name, member_type=member_type,
                organization=organization, start_date=start_date, end_date=end_date,
                status=status, notes=notes,
            )
        else:
            self.member_id = self.screen.member_service.create_member(
                full_name=full_name, member_type=member_type, organization=organization,
                start_date=start_date, end_date=end_date, notes=notes, status=status,
            )

        # حفظ الحقول الديناميكية
        for field_id, widget in self.dynamic_field_widgets.items():
            self.screen.dynamic_fields_service.set_member_value(self.member_id, field_id, widget.text().strip())

        self.accept()


# ==============================================================================
# 2. الشاشة الرئيسية لإدارة الأعضاء
# ==============================================================================
class MembersScreen(QWidget):
    def __init__(self, member_service, dynamic_fields_service, document_service, export_service):
        super().__init__()
        self.member_service = member_service
        self.dynamic_fields_service = dynamic_fields_service
        self.document_service = document_service
        self.export_service = export_service

        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self._build_ui()
        self.refresh_list()

    def refresh(self):
        self.refresh_list()

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(25, 20, 25, 20)
        main_layout.setSpacing(15)

        # ---------- شريط الأدوات والفلاتر العلوي ----------
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(12)

        filter_layout.addWidget(QLabel("النوع:"))
        self.type_filter = QComboBox()
        self.type_filter.setMinimumHeight(35)
        self.type_filter.setMinimumWidth(120)
        self.type_filter.addItem("الكل", None)
        for t in MEMBER_TYPES:
            self.type_filter.addItem(MEMBER_TYPE_LABELS_AR[t], t)
        self.type_filter.currentIndexChanged.connect(self.refresh_list)
        filter_layout.addWidget(self.type_filter)

        filter_layout.addWidget(QLabel("الحالة:"))
        self.status_filter = QComboBox()
        self.status_filter.setMinimumHeight(35)
        self.status_filter.setMinimumWidth(120)
        self.status_filter.addItem("الكل", None)
        for s in MEMBER_STATUSES:
            self.status_filter.addItem(MEMBER_STATUS_LABELS_AR[s], s)
        self.status_filter.currentIndexChanged.connect(self.refresh_list)
        filter_layout.addWidget(self.status_filter)

        self.search_input = QLineEdit()
        self.search_input.setMinimumHeight(35)
        self.search_input.setPlaceholderText("البحث السريع بالاسم المكتوب...")
        self.search_input.textChanged.connect(self.refresh_list)
        filter_layout.addWidget(self.search_input, 1)

        # الأزرار الرئيسية بالتأثيرات الفاخرة المقاومة للتمدد
        # زر التصدير الإجمالي القديم
        export_btn = QPushButton("تصدير Excel")
        export_btn.setFixedSize(110, 35)
        export_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; font-weight: bold; border-radius: 4px; }
            QPushButton:hover { background-color: #219653; }
        """)
        export_btn.clicked.connect(self._export_excel)
        filter_layout.addWidget(export_btn)

        # 👇 الزر الجديد: التصدير التفصيلي للحركات والماليات
        detailed_export_btn = QPushButton("📊 تقرير تفصيلي")
        detailed_export_btn.setFixedSize(120, 35)
        detailed_export_btn.setStyleSheet("""
            QPushButton { background-color: #8e44ad; color: white; font-weight: bold; border-radius: 4px; }
            QPushButton:hover { background-color: #732d91; }
        """)
        detailed_export_btn.clicked.connect(self._export_detailed_financial_excel)
        filter_layout.addWidget(detailed_export_btn)
        

        add_member_btn = QPushButton("+ عضو جديد")
        add_member_btn.setFixedSize(110, 35)
        add_member_btn.setStyleSheet("""
            QPushButton { background-color: #00a8ff; color: white; font-weight: bold; border-radius: 4px; }
            QPushButton:hover { background-color: #0097e6; }
        """)
        add_member_btn.clicked.connect(self._open_add_member_dialog)
        filter_layout.addWidget(add_member_btn)

        main_layout.addLayout(filter_layout)

        # ---------- جدول الأعضاء الرئيسي (كامل المساحة) ----------
        self.members_table = QTableWidget()
        self.members_table.setColumnCount(7)
        self.members_table.setHorizontalHeaderLabels([
            "الاسم الكامل", "النوع", "الجهة/المؤسسة", "تاريخ التعيين", "إجمالي المستحقات", "الحالة", "الإجراءات"
        ])
        self.members_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.members_table.setAlternatingRowColors(True)
        self.members_table.setStyleSheet("""
            QTableWidget {
                background-color: #2f3640;
                alternate-background-color: #1e272e;
                color: #f5f6fa;
                gridline-color: #718093;
                border: 1px solid #718093;
                border-radius: 8px;
            }
            QTableWidget::item {
                background-color: transparent;
                color: #f5f6fa;
                padding: 8px;
            }
            QTableWidget::item:selected {
                background-color: #00a8ff;
                color: white;
            }
            QHeaderView::section {
                background-color: #1e272e;
                color: #f5f6fa;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #718093;
            }
        """)
        main_layout.addWidget(self.members_table)


    def refresh_list(self):
        member_type = self.type_filter.currentData()
        status = self.status_filter.currentData()
        search = self.search_input.text().strip() or None

        self.members_table.setRowCount(0)
        members = self.member_service.list_members(member_type=member_type, status=status, search=search)

        for row_idx, m in enumerate(members):
            self.members_table.insertRow(row_idx)
            self.members_table.setRowHeight(row_idx, 45)

            # 0. الاسم الكامل
            self.members_table.setItem(row_idx, 0, QTableWidgetItem(m["full_name"]))
            # 1. النوع
            self.members_table.setItem(row_idx, 1, QTableWidgetItem(MEMBER_TYPE_LABELS_AR[m["member_type"]]))
            # 2. الجهة
            self.members_table.setItem(row_idx, 2, QTableWidgetItem(m["organization"] or "غير محدد"))
            # 3. تاريخ التعيين
            self.members_table.setItem(row_idx, 3, QTableWidgetItem(m["start_date"]))
            
            # 4. إجمالي المستحقات الفعلي (مجموع الصرف الفعلي والمكافآت)
            total_money = float(m.get("total_allowances", 0.0))
            money_item = QTableWidgetItem(f"{total_money:,.2f} ج.م")
            money_item.setForeground(QColor("#f1c40f"))  # لون ذهبي فخم للمبالغ المالية
            money_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.members_table.setItem(row_idx, 4, money_item)
            
            # 5. الحالة
            status_item = QTableWidgetItem(MEMBER_STATUS_LABELS_AR[m["status"]])
            if m["status"] == "Active":
                status_item.setForeground(QColor("#2ecc71"))
            else:
                status_item.setForeground(QColor("#e74c3c"))
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.members_table.setItem(row_idx, 5, status_item)

            # 6. عمود الإجراءات والتعديل
            actions_grid = QGridLayout()
            actions_grid.setContentsMargins(4, 2, 4, 2)
            actions_grid.setSpacing(6)

            # زر التعديل
            edit_btn = QPushButton("✏️")  
            edit_btn.setFixedSize(40, 28)  
            edit_btn.setToolTip("تعديل بيانات العضو")  
            edit_btn.setStyleSheet("""
                QPushButton { 
                    background-color: transparent; 
                    color: #00a8ff !important; 
                    font-weight: bold; 
                    font-size: 13px; 
                    border: 1px solid #00a8ff; 
                    border-radius: 4px; 
                }
                QPushButton:hover { 
                    background-color: #00a8ff; 
                    color: white !important; 
                }
            """)
            edit_btn.clicked.connect(lambda checked, mid=m["id"]: self._open_edit_member_dialog(mid))
            
            # زر الحذف
            delete_btn = QPushButton("🗑️")
            delete_btn.setFixedSize(40, 28)  
            delete_btn.setToolTip("حذف العضو")  
            delete_btn.setStyleSheet("""
                QPushButton { 
                    background-color: transparent; 
                    color: #e74c3c !important; 
                    font-weight: bold; 
                    font-size: 13px; 
                    border: 1px solid #e74c3c; 
                    border-radius: 4px; 
                }
                QPushButton:hover { 
                    background-color: #e74c3c; 
                    color: white !important; 
                }
            """)
            delete_btn.clicked.connect(lambda checked, mid=m["id"]: self._delete_member(mid))

            actions_grid.addWidget(edit_btn, 0, 0, Qt.AlignmentFlag.AlignCenter)
            actions_grid.addWidget(delete_btn, 0, 1, Qt.AlignmentFlag.AlignCenter)

            actions_grid.setColumnStretch(0, 1)
            actions_grid.setColumnStretch(1, 1)

            actions_widget = QWidget()
            actions_widget.setLayout(actions_grid)
            actions_widget.setStyleSheet("background-color: transparent;")
            self.members_table.setCellWidget(row_idx, 6, actions_widget)
            
        self.members_table.setColumnWidth(6, 110)

    # ------------------------------------------------------------- الأحداث الفرعية والربط البرمجي
    def _open_add_member_dialog(self):
        dialog = MemberFormDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_list()
            QMessageBox.information(self, "تم", "تم تسجيل العضو الجديد بنجاح.")

    def _open_edit_member_dialog(self, member_id):
        dialog = MemberFormDialog(self, member_id=member_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_list()
            QMessageBox.information(self, "تم", "تم تحديث بيانات العضو بنجاح.")

    def _delete_member(self, member_id):
        confirm = QMessageBox.question(
            self, "تأكيد الحذف", "هل أنت متأكد من حذف هذا العضو؟ لا يمكن التراجع عن هذا الإجراء.",
        )
        if confirm == QMessageBox.StandardButton.Yes:
            self.member_service.delete_member(member_id)
            self.refresh_list()
            QMessageBox.information(self, "تم", "تم حذف العضو من قاعدة البيانات.")

    def _export_excel(self):
        member_type = self.type_filter.currentData()
        status = self.status_filter.currentData()
        search = self.search_input.text().strip() or None
        
        # جلب الأعضاء شاملاً الحقول المالية المحسوبة من الجداول المختلفة
        members = self.member_service.list_members(member_type=member_type, status=status, search=search)
        if not members:
            QMessageBox.information(self, "تنبيه", "لا توجد بيانات متاحة لتصديرها حاليًا.")
            return

        # تجهيز النصوص والبيانات المالية بصيغة نصية مناسبة لملف الإكسل
        for m in members:
            m["member_type_label"] = MEMBER_TYPE_LABELS_AR[m["member_type"]]
            m["status_label"] = MEMBER_STATUS_LABELS_AR[m["status"]]
            
            # تحويل القيمة المالية لشكل نصي منسق داخل الإكسل
            total_money = float(m.get("total_allowances", 0.0))
            m["total_allowances_label"] = f"{total_money:,.2f} ج.م"

        file_path, _ = QFileDialog.getSaveFileName(self, "حفظ ملف Excel", "members_export.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return

        # إضافة العمود المالي الجديد "total_allowances_label" لترتيب العناوين
        headers = {
            "full_name": "الاسم",
            "member_type_label": "النوع",
            "organization": "الجهة",
            "start_date": "تاريخ التعيين",
            "total_allowances_label": "إجمالي المستحقات",  # العمود المالي الجديد
            "status_label": "الحالة",
        }
        
        self.export_service.export_to_excel(members, headers, file_path, sheet_title="Members")
        QMessageBox.information(self, "تم", f"تم التصدير بنجاح إلى:\n{file_path}")

    # 👇 الدالة الناقصة التي سببت الخطأ - تأكد من ضبط مسافاتها (Indentation) لتكون داخل الكلاس
    def _export_detailed_financial_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        member_type = self.type_filter.currentData()
        status = self.status_filter.currentData()
        search = self.search_input.text().strip() or None
        
        members = self.member_service.list_members(member_type=member_type, status=status, search=search)
        if not members:
            QMessageBox.information(self, "تنبيه", "لا توجد بيانات أعضاء لتصدير تقاريرهم.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "حفظ كشف الحركات المالي التفصيلي", "detailed_members_report.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return

        # إنشاء كتاب عمل إكسل جديد وتجهيز الشيت
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "كشف الحركات التفصيلي"
        ws.views.sheetView[0].rightToLeft = True # ضبط الشيت من اليمين لليسار عربي

        # ترويسة الجدول الرئيسية
        headers = ["القسم / بيان الحركة", "التفاصيل / التاريخ", "الحالة / الدور التنظيمي", "المبالغ المالية والبدلات"]
        ws.append(headers)

        # تجهيز الألوان والخطوط الفخمة للتنسيق
        header_fill = PatternFill(start_color="1B3A4B", end_color="1B3A4B", fill_type="solid") # أزرق غامق ملكي
        member_fill = PatternFill(start_color="D9E2EC", end_color="D9E2EC", fill_type="solid") # رمادي أزرق فاتح جداً للأعضاء
        section_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid") # تفتيح بسيط للأقسام الداخلية
        
        font_header = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
        font_member = Font(name="Segoe UI", size=12, bold=True, color="102A43")
        font_section = Font(name="Segoe UI", size=11, bold=True, color="334E68")
        font_data = Font(name="Segoe UI", size=10)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='BCCCDC'),
            right=Side(style='thin', color='BCCCDC'),
            top=Side(style='thin', color='BCCCDC'),
            bottom=Side(style='thin', color='BCCCDC')
        )

        # تنسيق صف الرؤوس الأول
        for col_num in range(1, 5):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = font_header
            cell.alignment = align_center

        current_row = 2

        for m in members:
            report = self.member_service.get_member_detailed_financial_report(m["id"])
            
            # 1. سطر اسم العضو الرئيسي (دمج 4 خلايا)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1)
            cell.value = f"👤 كشف حركات العضو: {m['full_name']} ({MEMBER_TYPE_LABELS_AR[m['member_type']]})"
            cell.fill = member_fill
            cell.font = font_member
            cell.alignment = align_right
            
            # تطبيق البوردر على الخلايا المدمجة لاسم العضو
            for c in range(1, 5):
                ws.cell(row=current_row, column=c).border = thin_border
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # --- أ) إضافة اللجان ---
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1)
            cell.value = "📋 [اللجان المشترك بها والمكافآت المعتمدة]"
            cell.fill = section_fill
            cell.font = font_section
            for c in range(1, 5): ws.cell(row=current_row, column=c).border = thin_border
            ws.row_dimensions[current_row].height = 22
            current_row += 1
            
            if report["committees"]:
                for comm in report["committees"]:
                    ws.cell(row=current_row, column=1, value=f"• لجنة: {comm['committee_name']}").font = font_data
                    ws.cell(row=current_row, column=2, value=f"بدل الحضور الافتراضي: {comm['session_rate']}").font = font_data
                    ws.cell(row=current_row, column=3, value=f"الدور: {comm['role'] or 'عضو'}").font = font_data
                    
                    money_cell = ws.cell(row=current_row, column=4, value=comm['bonus_amount'])
                    money_cell.font = font_data
                    money_cell.number_format = '#,##0.00" ج.م"' # تنسيق مالي احترافي في الإكسل
                    
                    for c in range(1, 5):
                        ws.cell(row=current_row, column=c).border = thin_border
                        if c > 1: ws.cell(row=current_row, column=c).alignment = align_center
                    ws.row_dimensions[current_row].height = 20
                    current_row += 1
            else:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                cell = ws.cell(row=current_row, column=1)
                cell.value = "   لا توجد لجان مسجلة لهذا العضو حالياً."
                cell.font = font_data
                for c in range(1, 5): ws.cell(row=current_row, column=c).border = thin_border
                current_row += 1

            # --- ب) إضافة حضور وغياب الجلسات والجمعيات ---
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1)
            cell.value = "📅 [حضور وغياب الجلسات والجمعيات العمومية]"
            cell.fill = section_fill
            cell.font = font_section
            for c in range(1, 5): ws.cell(row=current_row, column=c).border = thin_border
            ws.row_dimensions[current_row].height = 22
            current_row += 1
            
            if report["meetings"]:
                for meet in report["meetings"]:
                    status_text = "حضور" if meet["attended"] == 1 else "غياب"
                    type_label = "مجلس إدارة" if meet["session_type"] == "Meeting" else "لجنة فرعية"
                    
                    ws.cell(row=current_row, column=1, value=f"• {meet['meeting_name']} ({type_label})").font = font_data
                    ws.cell(row=current_row, column=2, value=f"التاريخ: {meet['meeting_date']}").font = font_data
                    ws.cell(row=current_row, column=3, value=status_text).font = font_data
                    
                    money_cell = ws.cell(row=current_row, column=4, value=meet['amount_paid'])
                    money_cell.font = font_data
                    money_cell.number_format = '#,##0.00" ج.م"'
                    
                    for c in range(1, 5):
                        ws.cell(row=current_row, column=c).border = thin_border
                        if c > 1: ws.cell(row=current_row, column=c).alignment = align_center
                    ws.row_dimensions[current_row].height = 20
                    current_row += 1
            else:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                cell = ws.cell(row=current_row, column=1)
                cell.value = "   لم يسجل للعضو أي حركة حضور أو غياب للجلسات."
                cell.font = font_data
                for c in range(1, 5): ws.cell(row=current_row, column=c).border = thin_border
                current_row += 1
                
            # سطر فارغ تماماً للفصل الأنيق
            current_row += 1

        # ضبط تلقائي لعرض الأعمدة بناءً على أطول نص
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 20)

        # حفظ الملف المنسق بنجاح
        wb.save(file_path)
        QMessageBox.information(self, "تم", f"تم تصدير الكشف المالي والإداري المنسق بنجاح إلى:\n{file_path}")