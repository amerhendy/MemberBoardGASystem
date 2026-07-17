# -*- coding: utf-8 -*-
"""
export_service.py
خدمة التصدير إلى Excel - Export Service
تستخدم openpyxl لتصدير أي قائمة بيانات (dicts) إلى ملف xlsx منسّق بشكل بسيط،
مع دعم اتجاه RTL للأعمدة العربية.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExportService:
    @staticmethod
    def export_to_excel(rows: list, headers: dict, file_path: str, sheet_title: str = "Report"):
        """
        rows: قائمة dicts (كل dict يمثل صف بيانات)
        headers: dict بصيغة {field_key: "عنوان العمود"} يحدد الأعمدة وترتيبها
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31] if sheet_title else "Report"
        ws.sheet_view.rightToLeft = True

        # صف العناوين
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        for col_idx, (key, label) in enumerate(headers.items(), start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # صفوف البيانات
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, key in enumerate(headers.keys(), start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))

        # عرض أعمدة تلقائي تقريبي
        for col_idx, key in enumerate(headers.keys(), start=1):
            max_len = max(
                [len(str(headers[key]))] + [len(str(r.get(key, ""))) for r in rows]
            ) if rows else len(str(headers[key]))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max_len + 4

        wb.save(file_path)
        return file_path
